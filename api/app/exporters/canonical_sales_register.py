"""
CSV exporter for canonical format sales_register.

Exports canonical v0.1 sales_register to CSV format.
Excel will happily open CSV; this is usually enough.
"""

import csv
from typing import Dict, Any, List, TextIO, Optional
import io
import logging

logger = logging.getLogger(__name__)


def sales_register_to_rows(doc: Dict[str, Any]) -> List[List[Any]]:
    """
    Convert canonical sales_register doc to flat rows for CSV/Excel export.
    
    Args:
        doc: Canonical format sales_register document
    
    Returns:
        List of rows (each row is a list of values)
    """
    rows: List[List[Any]] = []
    
    entries = doc.get("entries") or []
    
    for e in entries:
        party = e.get("party") or {}
        amounts = e.get("amounts") or {}
        tax_b = amounts.get("tax_breakup") or {}
        doc_spec = e.get("doc_specific") or {}
        
        rows.append([
            e.get("entry_id"),
            e.get("entry_number"),
            e.get("entry_date"),
            party.get("name"),
            party.get("gstin"),
            party.get("state_code"),
            amounts.get("taxable_value"),
            tax_b.get("cgst"),
            tax_b.get("sgst"),
            tax_b.get("igst"),
            tax_b.get("cess"),
            amounts.get("total"),
            doc_spec.get("reverse_charge"),
            doc_spec.get("invoice_type"),
            doc_spec.get("place_of_supply"),
        ])
    
    return rows


def export_sales_register_to_csv(doc: Dict[str, Any], file_path: str) -> None:
    """
    Write canonical sales_register to CSV file.
    
    Columns are chosen to be useful for CAs / reconciliation.
    
    Args:
        doc: Canonical format sales_register document
        file_path: Path to write CSV file
    """
    header = [
        "Entry ID",
        "Invoice No",
        "Invoice Date",
        "Customer Name",
        "Customer GSTIN",
        "Customer State Code",
        "Taxable Value",
        "CGST",
        "SGST",
        "IGST",
        "CESS",
        "Total Invoice Value",
        "Reverse Charge",
        "Invoice Type",
        "Place of Supply",
    ]
    
    rows = sales_register_to_rows(doc)
    
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)


def canonical_sales_register_to_csv(doc: Dict[str, Any]) -> str:
    """
    Export canonical format sales_register to CSV string (for API responses).
    
    Args:
        doc: Canonical format sales_register document
    
    Returns:
        CSV string with headers and rows
    """
    header = [
        "Entry ID",
        "Invoice No",
        "Invoice Date",
        "Customer Name",
        "Customer GSTIN",
        "Customer State Code",
        "Taxable Value",
        "CGST",
        "SGST",
        "IGST",
        "CESS",
        "Total Invoice Value",
        "Reverse Charge",
        "Invoice Type",
        "Place of Supply",
    ]
    
    rows = sales_register_to_rows(doc)
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(header)
    writer.writerows(rows)
    
    return output.getvalue()


def canonical_sales_register_to_excel(canonical: Dict[str, Any]) -> bytes:
    """
    Export canonical format sales_register to Excel (XLSX) format.
    
    Args:
        canonical: Canonical format sales_register
    
    Returns:
        Excel file as bytes (XLSX format)
    
    Note: Requires openpyxl library. Falls back to CSV if not available.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Register"
        
        # Headers
        headers = [
            "Entry ID",
            "Invoice No",
            "Invoice Date",
            "Customer Name",
            "Customer GSTIN",
            "Customer State Code",
            "Taxable Value",
            "CGST",
            "SGST",
            "IGST",
            "CESS",
            "Total Invoice Value",
            "Reverse Charge",
            "Invoice Type",
            "Place of Supply",
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        rows = sales_register_to_rows(canonical)
        
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = "#,##0.00"
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
        
    except ImportError:
        logger.warning("openpyxl not available, falling back to CSV format")
        # Fallback: return CSV as bytes
        csv_data = canonical_sales_register_to_csv(canonical)
        return csv_data.encode("utf-8")
    except Exception as e:
        logger.error(f"Error generating Excel file: {e}")
        # Fallback: return CSV as bytes
        csv_data = canonical_sales_register_to_csv(canonical)
        return csv_data.encode("utf-8")
